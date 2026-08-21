#!/usr/bin/env python3
"""
recalc.py — Recalculate Excel formulas using LibreOffice headless mode.

openpyxl writes formula strings but does NOT compute values. Excel will
recalculate on open, but for automated pipelines (validation, QC, handoff)
we need calculated values written back into the file. This script uses
LibreOffice headless conversion to force recalculation.

Usage:
    python scripts/recalc.py <file.xlsx> [timeout_seconds]

Requires:
    - LibreOffice (soffice in PATH)
        macOS:  brew install --cask libreoffice
        Linux:  apt install libreoffice  (or equivalent)
    - openpyxl (for post-recalc error scanning)
        pip install openpyxl

Returns JSON to stdout with:
    - status: "success" | "errors_found" | "failed"
    - total_formulas: count of formulas in file
    - total_errors: count of Excel error values
    - error_summary: {error_type: {count, locations}}
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path


EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def find_soffice() -> str | None:
    """Locate the LibreOffice binary."""
    for candidate in ("soffice", "libreoffice"):
        path = shutil.which(candidate)
        if path:
            return path
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if Path(mac_path).exists():
        return mac_path
    return None


def recalc_with_libreoffice(xlsx_path: Path, timeout: int) -> None:
    """Force formula recalculation by round-tripping through LibreOffice."""
    soffice = find_soffice()
    if soffice is None:
        raise RuntimeError(
            "LibreOffice not found. Install:\n"
            "  macOS: brew install --cask libreoffice\n"
            "  Linux: apt install libreoffice"
        )

    outdir = xlsx_path.parent
    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--calc",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(outdir),
            str(xlsx_path),
        ],
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed (exit {result.returncode}):\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )


def scan_errors(xlsx_path: Path) -> dict:
    """Scan all cells for formulas and Excel error values."""
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl") from e

    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)

    total_formulas = 0
    errors: dict[str, dict] = {}

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        for row in ws_f.iter_rows():
            for cell in row:
                try:
                    coord = cell.coordinate
                except AttributeError:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
                    computed = ws_v[coord].value
                    if isinstance(computed, str) and computed in EXCEL_ERRORS:
                        entry = errors.setdefault(
                            computed, {"count": 0, "locations": []}
                        )
                        entry["count"] += 1
                        if len(entry["locations"]) < 25:
                            entry["locations"].append(f"{sheet_name}!{coord}")

    total_errors = sum(e["count"] for e in errors.values())
    return {
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": errors,
    }


def main() -> int:
    if len(sys.argv) < 2:
        print(
            json.dumps(
                {
                    "status": "failed",
                    "error": "usage: python recalc.py <file.xlsx> [timeout_seconds]",
                }
            )
        )
        return 2

    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    timeout = int(sys.argv[2]) if len(sys.argv) >= 3 else 120

    if not xlsx_path.exists():
        print(json.dumps({"status": "failed", "error": f"File not found: {xlsx_path}"}))
        return 2

    try:
        recalc_with_libreoffice(xlsx_path, timeout)
    except Exception as e:
        print(json.dumps({"status": "failed", "error": str(e)}))
        return 1

    try:
        scan = scan_errors(xlsx_path)
    except Exception as e:
        print(json.dumps({"status": "failed", "error": f"scan failed: {e}"}))
        return 1

    status = "errors_found" if scan["total_errors"] > 0 else "success"
    print(json.dumps({"status": status, **scan}, indent=2))
    return 0 if status == "success" else 1


if __name__ == "__main__":
    sys.exit(main())
